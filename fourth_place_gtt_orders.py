import openpyxl

from datetime import date

import helper
from config import WORKBOOK_FOR_SST, WORKBOOK_TRACKING_SHEET


def init():
    #This login is only required when runing independently
    #helper.login_kite_in_browser()

    #Read Tracking sheet and place orders
    # Loading the workbook for SST
    wb = openpyxl.load_workbook(WORKBOOK_FOR_SST)

    # Loading the Tracking sheet from Workbook
    sh2 = wb[WORKBOOK_TRACKING_SHEET]
    #storing the total number of row count in row variable
    row = sh2.max_row

    helper.delete_all_gtt_orders()
    # Iterating all the rows in sst.xlsx
    d = date.today()
    for i in range(1, row):
        sym = sh2.cell(i + 1, 1).value
        diff_20DH_CMP = int(sh2.cell(i + 1, 4).value)
        if 0<= int(diff_20DH_CMP) <= 5:
            # If the GTT Order Price column in excel is None then the 20 Day High is the GTT Price
            gtt_order_price = sh2.cell(i + 1, 3).value
            # Updating the 20 DH in the GTT order price column
            sh2.cell(i + 1, 6).value = gtt_order_price
            sh2.cell(i + 1, 5).value = d
            # Placing a buy order and this is a fresh scenario
            helper.place_buy_order( sym, gtt_order_price)
            print(f'Buy order placed for {sym}')
    #Saving the final workbook
    wb.save(WORKBOOK_FOR_SST)


def main():
    init()
main()