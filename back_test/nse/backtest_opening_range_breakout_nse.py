import os

import backtrader as bt

from back_test.nse.get_nse_data import get_historical_data
from config import *
import helper
import openpyxl
import pandas as pd
from pathlib import Path




class TestSSTStrategy(bt.Strategy):

    def calculated_dyn_per_away(self,number_of_buy):
        if number_of_buy == 1 :
            return 20
        elif number_of_buy == 2 :
            return 30
        elif number_of_buy == 3:
            return 40
        else :
            return 40

    def log(self, txt, dt=None):
        ''' Logging function for this strategy'''
        dt = dt or self.datas[0].datetime.date(0)
        print('%s, %s' % (dt.isoformat(), txt))

    #
    def __init__(self):
        self.the_highest_high_20 = bt.ind.Highest(self.data.high, period=20)
        self.the_lowest_low_20 = bt.ind.Lowest(self.data.low, period=20)
        self.datalow = self.datas[0].low # Resolved from issue solution -https://community.backtrader.com/topic/2411/trying-to-use-binance-data/2
        self.datahigh = self.datas[0].high

        # To keep track of pending orders and some init variables
        self.order = None
        self.gtt_price = 0.0
        self.started_tracking= False
        self.buy_price = 0.0
        self.target_price = 0.0
        self.buy_orders = []
        self.buy_avg = 0.0
        self.portfolio= 4000

    def next(self):
        #Module to start Tracking the coin
        if not self.started_tracking and self.datalow[0] <=self.the_lowest_low_20 :
            #If previous order exists then start tracking iff the current price is less then 15% of average buying price
            if len(self.buy_orders) > 0:
                #self.log('Previous order exists and current avg buy price is %.2f' % self.buy_avg)
                #Start New tracking iff iff the current price is less then 15% of average buying price
                per_dif = (self.buy_avg - (self.datahigh[0] + self.datalow[0])/2)/self.buy_avg * 100
                dyn_per_away_to_be_considered = self.calculated_dyn_per_away(number_of_buy=len(self.buy_orders))
                self.log('Required Dynamic Percentage away calculated is  %.2f' % dyn_per_away_to_be_considered)
                self.log('Actual Percentage away calculated is  %.2f' % per_dif)
                if per_dif >=dyn_per_away_to_be_considered:
                    self.log(f'Started Tracking the coin as it previous buying avg is less than {dyn_per_away_to_be_considered} per %.2f' % self.datalow[0])

                    # Update the GTT price as 20 DH
                    self.gtt_price = self.the_highest_high_20
                    self.started_tracking = True
            else:
                self.log('Start Tracking the coin, %.2f' % self.datalow[0])
                #Update the GTT price as 20 DH
                self.gtt_price = self.the_highest_high_20
                self.started_tracking = True


        #Place a buy order if todays high is greater than gtt_price and if the position is being tracked in Tracking sheet
        if self.started_tracking and self.datahigh[0] >=self.gtt_price and self.gtt_price != 0.0:
            self.buy_price = self.gtt_price + (self.gtt_price* 0.001)
            #self.buy_price = self.gtt_price
            self.log('New Buy Created, %.2f' % self.buy_price)
            #self.target_price = self.buy_price + self.buy_price*self.take_profit
            self.buy(size=1)

            #Reducing portfolio on each buy
            self.portfolio = self.portfolio - self.buy_price
            self.log('Portfolio after buy, %.2f' % self.portfolio)
            #Remove from tracking sheet once the buy position is taken
            self.started_tracking = False

            #Push a buy order in buy_orders list
            order = {'buy_price' :self.buy_price }
            self.buy_orders.append(order)

            total_orders = len(self.buy_orders)
            sum_buy_orders = 0.0
            for order in self.buy_orders:
                sum_buy_orders = sum_buy_orders +order.get('buy_price')
            self.buy_avg = sum_buy_orders/total_orders

        # Update gtt_price if the new high is made and position is not taken
        if self.started_tracking and self.gtt_price > 0.0 and not self.position :
            if self.gtt_price != self.the_highest_high_20:
                gtt_price= self.the_highest_high_20


        #Condition for sell
        if self.position:
            target_per = 0
            #Calculating dynamic target :
            if len(self.buy_orders) == 1 :
                target_per = 10
            elif len(self.buy_orders) == 2:
                target_per = 8
            elif len(self.buy_orders) == 3 :
                target_per = 5
            elif len(self.buy_orders) > 3 :
                target_per = 5
            self.target_price = self.buy_avg + (self.buy_avg *target_per)/100

            if self.datahigh[0] >=self.target_price and self.target_price > 0.0 :
                self.log('Average buying price for this order is, %.2f' % self.buy_avg)
                self.log('Target percentsge  for this sell is, %.2f' % target_per)
                self.log('Target price  for this sell is, %.2f' % self.target_price)
                self.log('Total selling quantity, %.2f' % len(self.buy_orders))
                self.log('SELL executed at , %.2f' % self.target_price)
                profit=(self.target_price - self.buy_avg) *len(self.buy_orders)
                self.log('Profit CREATE, %.2f' % profit)
                self.portfolio = self.portfolio+ self.target_price *len(self.buy_orders)
                self.log('Portfolio Value after sell, %.2f' % self.portfolio)
                print('=============================================================================================')
                self.order = self.sell()

                # Reset the init variables
                self.buy_orders = []
                self.gtt_price = 0.0
                self.started_tracking = False
                self.buy_price = 0.0
                self.buy_avg = 0.0

#===================================================Boiler Plate Code ====================================
if __name__ == '__main__':
    parent = os.path.dirname(os.getcwd())
    #print(os.path.dirname(parent))
    # Loading the workbook for SST
    wb = openpyxl.load_workbook(os.path.dirname(parent)+"\\"+WORKBOOK_FOR_SST)

    # Loading the Main sheet from Workbook
    sh1 = wb[WORKBOOK_BACKTEST_SHEET]

    #storing the total number of row count in row variable
    row = sh1.max_row

    # Iterating all the rows in sst.xlsx
    for i in range(1, row):
        get_historical_data(sh1.cell(i + 1, 1).value,1000)
        cerebro = bt.Cerebro()
        #Updating the default cash with broker
        cerebro.broker.setcash(10000.0)

        # Create a Data Feed
        data = bt.feeds.GenericCSVData(
            dataname='nse.csv',
            dtformat=('%b/%d/%Y'),
            #dtformat=('%YYYY-%mm-%dd'),
            datetime= 15,
            timestamp = 0,
            high = 5,
            low = 6,
            open = 4,
            close = 8,
            volume = 10,
            #timeframe=bt.TimeFrame.Days,
            #compression=1
            )

        cerebro.adddata(data)
        cerebro.addstrategy(TestSSTStrategy)

        print('Starting Portfolio Value: %.2f' % cerebro.broker.getvalue())
        print('=============================================================================================')
        print('=============================================================================================')
        cerebro.run()
        print('=============================================================================================')
        print('=============================================================================================')
        print('Final Portfolio Value: %.2f' % cerebro.broker.getvalue())

        #Uncomment this if you want plots
        #cerebro.plot()


