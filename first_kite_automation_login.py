import csv
import os
import time

import openpyxl
from selenium.webdriver.common.by import By

import helper
import kite_automation_config
from config import WORKBOOK_FOR_SST, WORKBOOK_INVESTMENT_SHEET, WORKBOOK_TRACKING_SHEET

#Login in Kite Browser
helper.login_kite_in_browser()

#Getting driver from helper
driver = helper.get_driver()

time.sleep(kite_automation_config.DELAY_IN_SEC)
element_holdings_menu = driver.find_element(By.XPATH,kite_automation_config.KITE_HOLDINGS_MENU_XPATH)
if element_holdings_menu.is_displayed() :
    element_holdings_menu.click()

time.sleep(kite_automation_config.DELAY_IN_SEC)
element_holdings_download_menu = driver.find_element(By.XPATH,kite_automation_config.KITE_HOLDINGS_DOWNLOAD_BTN_XPATH)
if element_holdings_download_menu.is_displayed() :
    element_holdings_download_menu.click()

time.sleep(kite_automation_config.DELAY_IN_SEC+5)
#Reading the holdings.csv file
# Loading the workbook for SST
wb = openpyxl.load_workbook(WORKBOOK_FOR_SST)

# Loading the Tracking sheet from Workbook
sh4 = wb[WORKBOOK_INVESTMENT_SHEET]
holdings_file = r'C:\Users\Rohit Pant\Downloads\holdings.csv'
with open(holdings_file) as file :
    reader = csv.reader(file)
    i = 1
    for row in reader:
        helper.sym_where_we_have_position.add(row[0])
        sh4.cell(i , 1).value = row[0]
        sh4.cell(i , 2).value = row[1]
        sh4.cell(i , 3).value = row[2]
        sh4.cell(i , 4).value = row[3]
        sh4.cell(i , 5).value = row[4]
        sh4.cell(i , 6).value = row[5]
        sh4.cell(i , 7).value = row[6]
        sh4.cell(i , 8).value = row[7]
        i=i+1
os.remove(holdings_file)
#driver.close()
#4. Delete rows in Tracking sheet where we already have positions
#wb = openpyxl.load_workbook(WORKBOOK_FOR_SST)
sh2 = wb[WORKBOOK_TRACKING_SHEET]
# Loop to delete rows in Tracking sheet
row2 = sh2.max_row
for i in range(1, row2):
    sym = sh2.cell(i + 1, 1).value
    if sym in helper.sym_where_we_have_position:
        sh2.delete_rows(idx=i + 1)
        print(f'Position in {sym} therefore deleting it from Tracking sheet')


wb.save(WORKBOOK_FOR_SST)


#driver.quit()