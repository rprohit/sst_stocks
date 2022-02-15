from openpyxl.styles import PatternFill

from config import *
import helper
import openpyxl
import pandas as pd
#Option to display all rows
#pd.set_option('display.max_rows',None)
pd.set_option('display.max_columns',None)


duration=20

def init():
    new_symbols_for_tracking = []

    # Loading the workbook for SST
    wb = openpyxl.load_workbook(WORKBOOK_FOR_SST)

    # Loading the Main sheet from Workbook
    sh1 = wb[WORKBOOK_MAIN_SHEET]

    #storing the total number of row count in row variable
    row = sh1.max_row

    #storing the total number of cloumn count in col variable
    col = sh1.max_column

    # Iterating all the rows in sst.xlsx
    for i in range(1, row):
        # Skipping the header using + 1 and reading the first column of each row in variable sym
        sym = sh1.cell(i + 1, 1).value

        # Passing the symbol to get the last 20 day data which is a list of list
        candles = helper.get_historic_data_by_symbol_days(sym, 40)
        df = pd.DataFrame(candles)

        print(f'At Row {i + 1} For symbol {sym}')

        # Below will add the row which represents last 20 day low record and the date
        # Behavior could be achived without rolling as well
        df['low_20'] = df['Low'].rolling(20).min()
        df['low_3'] = df['Low'].rolling(3).min()

        # Below will add the row which represents last 20 day High record and the date
        df['high_20_day'] = df['High'].rolling(20).max()

        #Fethcing the last row
        last_row = df.iloc[-1]
        low_20_day = float(last_row['low_20'])
        for current in range(0, len(df.index)):
            if float(df['Low'][current]) == low_20_day:
                #print(f'{low_20_day} and {df["Low"][current]}')
                df['low_20_day_date'] =df.index[current]
                #break
        last_row = df.iloc[-1]
        #fetching the low date
        low_20_day_date = str(last_row['low_20_day_date'])

        #fetching the high
        high_20_day = float(last_row['high_20_day'])

        sh1.cell(i + 1, 3).value = float(low_20_day)
        sh1.cell(i + 1, 4).value = str(low_20_day_date)

        # TODO Code to get last 20 day high from today and last 20 day high from t-1 day
        if sh1.cell(i + 1, 5).value is None:
            sh1.cell(i + 1, 5).value = float(high_20_day)
        else:
            sh1.cell(i + 1, 6).value = float(high_20_day)

        # Below will query and fetch the lastest spot price of the symbol
        spot = pd.DataFrame(helper.get_historic_data_by_symbol_days(sym, 7)).iloc[-1]

        # storing the spot price in excel
        sh1.cell(i + 1, 2).value = spot['Close']

        lowest_from_3_day = float(last_row['low_3'])
        sh1.cell(i + 1, 8).value = float(lowest_from_3_day)

        # Missed the opportunity in last 3 days
        if lowest_from_3_day == low_20_day:
            sh1.cell(i + 1, 9).value = 'Yes'
            new_symbols_for_tracking.append(sym)

        # Populate Away %
        away_per = ((float(high_20_day) - float(spot['Close'])) / float(spot['Close'])) * 100
        sh1.cell(i + 1, 7).value = int(away_per)

        # Adding a blue color background if the difference between the trigger price and current price is less thn 5
        if int(away_per) <= 5:
            sh1.cell(i + 1, 7).fill = PatternFill(start_color="d1d2ef", end_color="d1d2ef", fill_type="solid")
        else:
            sh1.cell(i + 1, 7).fill = PatternFill(fill_type=None)



    # Add these new symbols in Tracking sheet if already not present
    # Loading the Tracking sheet from Workbook
    sh2 = wb[WORKBOOK_TRACKING_SHEET]
    # storing the total number of row count in row variable
    row = sh2.max_row

    # Iterating all the rows in sst.xlsx in Tracking sheet and fetch the symbols which are already being tracked
    previous_sym_already_tracked = []
    for i in range(1, row):
        previous_sym_already_tracked.append(sh2.cell(i + 1, 1).value)


    #Iterating all the rows where position is already taken and the next averaage will be done at 40% loss
    position_already_taken = []
    sh3 = wb[WORKBOOK_INVESTMENT_SHEET]
    #storing the total number of row count in row variable
    row3 = sh3.max_row
    for i in range(1, row):
        position_already_taken.append(sh3.cell(i + 1, 1).value)
    #Apending new symbols in tracking sheet
    #TODO Need to add 20% loss login for second average
    sym_not_in_tracking_sheet = set(new_symbols_for_tracking) - set(previous_sym_already_tracked)- set(position_already_taken)
    if len(sym_not_in_tracking_sheet) > 0 :
        sym_not_in_tracking_sheet_list = list(sym_not_in_tracking_sheet)
        for i in range(0,len(sym_not_in_tracking_sheet_list)):
            sh2.cell(row+ i + 1, 1).value = sym_not_in_tracking_sheet_list[i]


    #Saving the final workbook
    wb.save(WORKBOOK_FOR_SST)

def main():
    init()

main()