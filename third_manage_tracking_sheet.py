import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill

import helper
from config import *


def init():

    # Loading the workbook for SST
    wb = openpyxl.load_workbook(WORKBOOK_FOR_SST)

    # Loading the Tracking sheet from Workbook
    sh2 = wb[WORKBOOK_TRACKING_SHEET]

    #storing the total number of row count in row variable
    row = sh2.max_row

    # Iterating all the rows in sst.xlsx
    for i in range(1, row):
        # Skipping the header using + 1 and reading the first column of each row in variable sym
        sym = sh2.cell(i + 1, 1).value
        spot = pd.DataFrame(helper.get_historic_data_by_symbol_days(sym, 7)).iloc[-1]
        #storing the spot price in excel

        #Fetching the closing price
        spot_price = spot['Close']
        sh2.cell(i + 1, 2).value = spot_price

        #Passing the symbol to get the last 20 day data which is a list of list
        candles = helper.get_historic_data_by_symbol_days(sym, 40)
        df = pd.DataFrame(candles)
        print(f'At Row {i + 1} For symbol {sym}')

        #Below will return last 20 day high record
        #highest_row = max(candles, key=lambda x: x[2])

        #fetching the high
        df['high_20_day'] = df['High'].rolling(20).max()
        last_row = df.iloc[-1:]
        high_20_day = float(last_row['high_20_day'])
        sh2.cell(i + 1, 3).value = high_20_day

        diff_20DH_CMP = ((float(high_20_day) - float(spot_price))/float(spot_price) ) * 100

        sh2.cell(i + 1, 4).value = int(diff_20DH_CMP)
        #Adding a blue color background if the difference between the trigger price and current price is less thn 5
        if 0<= int(diff_20DH_CMP) <= 5:
            sh2.cell(i + 1, 4).fill =PatternFill(start_color="d1d2ef", end_color="d1d2ef", fill_type="solid")
        else:
            sh2.cell(i + 1, 4).fill = PatternFill(fill_type=None)


        #Logic to update GTT price
        update_gtt_order = ''
        gtt_order_price = sh2.cell(i + 1, 6).value
        if gtt_order_price is not None and gtt_order_price != high_20_day:
            update_gtt_order = 'Y'
        else:
            update_gtt_order = ''
        sh2.cell(i + 1, 8).value = update_gtt_order
    #Saving the workbook
    wb.save(WORKBOOK_FOR_SST)
def main():
    init()

main()